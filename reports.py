"""
reports.py
──────────
وحدة التقارير — تصدير البيانات كـ Excel أو CSV
يستخدم openpyxl لإنشاء ملفات Excel منسقة احترافياً
"""

import pandas as pd
import io
from datetime import date


def _style_excel_sheet(worksheet, df: pd.DataFrame, header_color: str = "1E293B"):
    """
    دالة مساعدة لتنسيق صفحة Excel:
    - تلوين الرأس
    - ضبط عرض الأعمدة تلقائياً
    - تنسيق الأرقام
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            bottom=Side(style='thin', color='CBD5E1')
        )

        # تنسيق صف الرأس
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # ضبط عرض الأعمدة تلقائياً
        for col_num, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            max_length = max(len(str(column)), 10)
            for row in worksheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = thin_border
            worksheet.column_dimensions[col_letter].width = min(max_length + 4, 35)

        # تجميد الصف الأول
        worksheet.freeze_panes = "A2"

    except ImportError:
        pass  # إذا لم يتوفر openpyxl للتنسيق، ننتج الملف بدون تنسيق


def export_inventory_excel(df: pd.DataFrame) -> bytes:
    """
    تصدير بيانات المخزن كملف Excel منسق
    Returns: bytes للملف — يُستخدم في st.download_button
    """
    # إعادة تسمية الأعمدة للعربية
    rename_map = {
        'id': 'الرقم',
        'brand': 'الماركة',
        'model': 'الموديل',
        'color': 'اللون',
        'imei': 'IMEI / Serial',
        'purchase_price': 'سعر الشراء (ج.م)',
        'expected_sale_price': 'سعر البيع المتوقع (ج.م)',
        'status': 'الحالة',
        'entry_date': 'تاريخ الإدخال',
        'days_in_stock': 'أيام في المخزن',
    }
    export_df = df.rename(columns=rename_map)

    # إضافة عمود الربح المتوقع
    if 'سعر الشراء (ج.م)' in export_df.columns and 'سعر البيع المتوقع (ج.م)' in export_df.columns:
        export_df['الربح المتوقع (ج.م)'] = (
            export_df['سعر البيع المتوقع (ج.م)'] - export_df['سعر الشراء (ج.م)']
        )

    # إنشاء الملف في الذاكرة
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='المخزن')

        # تنسيق الصفحة
        ws = writer.sheets['المخزن']
        _style_excel_sheet(ws, export_df, header_color="0E4D8C")

        # صفحة ثانية: ملخص
        summary_data = {
            'البيان': ['إجمالي الأجهزة', 'متاح للبيع', 'تم البيع', 'تاريخ التقرير'],
            'القيمة': [
                len(df),
                len(df[df['status'] == 'Available']) if 'status' in df.columns else 0,
                len(df[df['status'] == 'Sold']) if 'status' in df.columns else 0,
                str(date.today())
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name='الملخص')
        _style_excel_sheet(writer.sheets['الملخص'], pd.DataFrame(summary_data), header_color="155E21")

    return buffer.getvalue()


def export_sales_excel(df: pd.DataFrame) -> bytes:
    """
    تصدير بيانات المبيعات كملف Excel منسق مع صفحة ملخص
    """
    rename_map = {
        'sale_id': 'رقم البيع',
        'brand': 'الماركة',
        'model': 'الموديل',
        'imei': 'IMEI / Serial',
        'actual_sale_price': 'سعر البيع الفعلي (ج.م)',
        'profit': 'الربح (ج.م)',
        'customer_name': 'اسم العميل',
        'customer_phone': 'رقم الهاتف',
        'sale_date': 'تاريخ البيع',
    }
    export_df = df.rename(columns=rename_map)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='المبيعات')

        ws = writer.sheets['المبيعات']
        _style_excel_sheet(ws, export_df, header_color="7C3AED")

        # صفحة الملخص المالي
        total_revenue = df['actual_sale_price'].sum() if 'actual_sale_price' in df.columns else 0
        total_profit  = df['profit'].sum() if 'profit' in df.columns else 0
        avg_profit    = df['profit'].mean() if 'profit' in df.columns else 0

        summary_data = {
            'البيان': [
                'إجمالي عمليات البيع',
                'إجمالي الإيرادات (ج.م)',
                'صافي الربح (ج.م)',
                'متوسط الربح لكل صفقة (ج.م)',
                'تاريخ التقرير'
            ],
            'القيمة': [
                len(df),
                f"{total_revenue:,.2f}",
                f"{total_profit:,.2f}",
                f"{avg_profit:,.2f}",
                str(date.today())
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name='الملخص المالي')
        _style_excel_sheet(
            writer.sheets['الملخص المالي'],
            pd.DataFrame(summary_data),
            header_color="0E4D8C"
        )

    return buffer.getvalue()
